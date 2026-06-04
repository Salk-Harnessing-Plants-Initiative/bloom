#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "psycopg[binary]>=3.2",
#   "python-dotenv>=1.0",
# ]
# ///

"""
backfill_gene_candidates_from_excel.py — load gene candidates + their progress
log history from the HPI Gene Candidate List Excel into the database.

One-time data cleanup, NOT a schema migration. Lives under scripts/ so the
deploy workflow ignores it; run by hand once per environment.

The Excel sheet has 12 columns; we map them as:

    Gene ID        → gene_candidates.gene (PK)
    Category       → gene_candidates.category
    Status         → gene_candidates.status
    Subcategory    → experiment_progress_logs row, message = "Subcategory: ..."
    Dec 2024 Upd.  → experiment_progress_logs row, timestamp = 2024-12-15
    Jun 2025 Upd.  → experiment_progress_logs row, timestamp = 2025-06-15
                     (including the literal "No Updates" string)
    Discovery Scientist + Collaborating Scientist
                   → not mapped; emails are unknown so user_email is left NULL

USAGE
-----
Local against a dev DB:

    POSTGRES_DSN="postgres://postgres:postgres@127.0.0.1:5433/postgres" \\
      uv run scripts/backfill_gene_candidates_from_excel.py \\
      --excel-path ~/Downloads/HPI\\ Gene\\ Candidate\\ List\\ June\\ 2025.xlsx \\
      --dry-run

Against staging (run from inside the staging supanet, like render_plate_videos.py):

    cd /data/bloom/staging
    docker run --rm --network bloom_v2_staging_supanet \\
      -v "$PWD:/work" -v "$HOME/Downloads:/data" -w /work \\
      --env-file .env.staging \\
      -e POSTGRES_DSN="postgres://supabase_admin:$(grep -E '^POSTGRES_PASSWORD=' .env.staging | cut -d= -f2-)@db-prod:5432/postgres" \\
      ghcr.io/astral-sh/uv:python3.11-bookworm-slim \\
      uv run scripts/backfill_gene_candidates_from_excel.py \\
      --excel-path "/data/HPI Gene Candidate List June 2025.xlsx" \\
      --dry-run

Against prod: swap staging → production / bloom_v2_prod_supanet / .env.prod.

DESIGN NOTES
------------
- gene_candidates: INSERT ... ON CONFLICT (gene) DO NOTHING. Existing rows
  are left untouched (Excel data is treated as the cold-start seed, not the
  source of truth for already-curated candidates).
- experiment_progress_logs: each new log carries a `tags` JSONB entry
  {"label": "backfill-2025-06", "color": "#84cc16"} so re-runs can detect
  already-imported rows by (gene, message, timestamp) and skip them. The
  composite uniqueness check is a SELECT-then-INSERT pattern — fine at this
  scale (~352 rows × ≤3 logs each).
- All logs are inserted with user_email = NULL because the Excel only carries
  scientist names, not email addresses. The UI already handles NULL senders.
"""

import argparse
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
import psycopg
from psycopg.types.json import Json
from dotenv import load_dotenv

BACKFILL_TAG = {"label": "backfill-2025-06", "color": "#84cc16"}
DEC_2024_TS = datetime(2024, 12, 15, tzinfo=timezone.utc).isoformat()
JUN_2025_TS = datetime(2025, 6, 15, tzinfo=timezone.utc).isoformat()
# Used for the synthetic "Subcategory" log so it has a stable timestamp
# distinct from the real progress note timestamps.
SUBCAT_TS = datetime(2024, 12, 1, tzinfo=timezone.utc).isoformat()

# The gene_candidates_status_check constraint only allows these five values.
# The Excel uses a different vocabulary; map each free-form Excel status to
# the closest constraint-valid value. Unknown values fall back to
# 'suspected' (the lowest-confidence bucket — "we noticed this gene but
# haven't confirmed anything") and a warning is logged.
STATUS_MAP = {
    "Approved by TC": "in-translation",
    "Deprioritized": "stopped",
    "In progress": "under-investigation",
}
FALLBACK_STATUS = "suspected"


def normalize(v) -> str | None:
    """Excel cells come back as int/float/str/None — coerce to a trimmed str
    or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(
        description="Backfill gene_candidates + experiment_progress_logs from Excel"
    )
    ap.add_argument(
        "--excel-path",
        type=Path,
        required=True,
        help="Path to the HPI Gene Candidate List xlsx",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Roll back the transaction at the end — counts are still reported",
    )
    ap.add_argument(
        "-v",
        "--verbose",
        action="store_true",
    )
    args = ap.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    load_dotenv()
    dsn = os.environ.get("POSTGRES_DSN")
    if not dsn:
        sys.exit("Set POSTGRES_DSN env var (postgres://user:pw@host:port/db)")

    if not args.excel_path.exists():
        sys.exit(f"Excel file not found: {args.excel_path}")

    wb = load_workbook(args.excel_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        sys.exit("Workbook has no active sheet")
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    logging.info("loaded %d Excel rows (excluding header)", len(rows))

    n_candidates_inserted = 0
    n_candidates_skipped = 0
    n_logs_inserted = 0
    n_logs_skipped = 0
    n_rows_no_gene = 0

    conn = psycopg.connect(dsn, autocommit=False)
    try:
        with conn.cursor() as cur:
            for row in rows:
                cells = list(row) + [None] * (12 - len(row))
                (
                    _gene_id_num,
                    _gene_name,
                    category,
                    subcategory,
                    status,
                    _discovery_scientist,
                    _collaborating_scientist,
                    _ortholog_analysis,
                    _gene_nickname,
                    gene_identifier,
                    dec_updates,
                    jun_updates,
                ) = cells[:12]

                gene = normalize(gene_identifier)
                if not gene:
                    n_rows_no_gene += 1
                    continue

                # ─── genes row first (FK target for gene_candidates.gene) ───
                # The genes table is the FK referent. Insert a minimal row
                # (just gene_id) so the candidate FK validates; richer fields
                # like reference_id / symbol can be filled in later by ingest.
                cur.execute(
                    """
                    INSERT INTO public.genes (gene_id)
                    VALUES (%s)
                    ON CONFLICT (gene_id) DO NOTHING
                    """,
                    (gene,),
                )

                # ─── gene_candidates upsert ──────────────────────────────────
                status_raw = normalize(status)
                if status_raw is None:
                    status_db = FALLBACK_STATUS
                elif status_raw in STATUS_MAP:
                    status_db = STATUS_MAP[status_raw]
                else:
                    logging.warning(
                        "unknown status %r on gene %s; using %s",
                        status_raw, gene, FALLBACK_STATUS,
                    )
                    status_db = FALLBACK_STATUS

                cur.execute(
                    """
                    INSERT INTO public.gene_candidates (gene, category, status)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (gene) DO NOTHING
                    """,
                    (
                        gene,
                        normalize(category),
                        status_db,
                    ),
                )
                if cur.rowcount > 0:
                    n_candidates_inserted += 1
                else:
                    n_candidates_skipped += 1

                # ─── progress logs (up to three per gene) ────────────────────
                pending = []
                sub = normalize(subcategory)
                if sub:
                    pending.append((f"Subcategory: {sub}", SUBCAT_TS))
                dec = normalize(dec_updates)
                if dec:
                    pending.append((dec, DEC_2024_TS))
                jun = normalize(jun_updates)
                if jun:
                    pending.append((jun, JUN_2025_TS))

                for msg, ts in pending:
                    cur.execute(
                        """
                        SELECT 1 FROM public.experiment_progress_logs
                        WHERE gene = %s
                          AND message = %s
                          AND timestamp = %s
                        LIMIT 1
                        """,
                        (gene, msg, ts),
                    )
                    if cur.fetchone():
                        n_logs_skipped += 1
                        continue
                    cur.execute(
                        """
                        INSERT INTO public.experiment_progress_logs
                            (gene, message, timestamp, user_email, tags, links, images)
                        VALUES (%s, %s, %s, NULL, %s::jsonb, '[]'::jsonb, '[]'::jsonb)
                        """,
                        (gene, msg, ts, Json([BACKFILL_TAG])),
                    )
                    n_logs_inserted += 1

        if args.dry_run:
            conn.rollback()
            logging.info("DRY RUN — transaction rolled back")
        else:
            conn.commit()
    finally:
        conn.close()

    print()
    print(f"rows in sheet:                              {len(rows)}")
    print(f"rows without a Gene ID (skipped):           {n_rows_no_gene}")
    print(f"gene_candidates inserted:                   {n_candidates_inserted}")
    print(f"gene_candidates skipped (already existed):  {n_candidates_skipped}")
    print(f"experiment_progress_logs inserted:          {n_logs_inserted}")
    print(f"experiment_progress_logs skipped (dup):     {n_logs_skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
