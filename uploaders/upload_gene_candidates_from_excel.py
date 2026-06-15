#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "supabase>=2.0",
#     "openpyxl>=3.1",
# ]
# ///

"""Upload gene candidates + their progress logs from the HPI Excel file
into bloom via the Supabase REST API (PostgREST).

Uses the same HTTPS/REST path as the cyl scanners and graviscan — no
SSH, no docker network access needed. Authenticates with the
SERVICE_ROLE_KEY so it bypasses RLS for the write.

USAGE
-----

Staging:

    BLOOM_URL="https://staging.bloom.salk.edu:8443/api" \\
    SERVICE_ROLE_KEY="$(gh secret get STAGING_SERVICE_ROLE_KEY \\
                       -R Salk-Harnessing-Plants-Initiative/bloom)" \\
      uv run uploaders/upload_gene_candidates_from_excel.py \\
      --excel-path ~/Downloads/HPI\\ Gene\\ Candidate\\ List\\ June\\ 2025.xlsx \\
      --dry-run

Prod: swap BLOOM_URL to https://bloom.salk.edu/api and SERVICE_ROLE_KEY
to PROD_SERVICE_ROLE_KEY.

(`gh secret get` requires gh >= 2.45. If your gh is older, copy the
value from the GitHub Secrets UI.)

EXCEL SHAPE
-----------
12 columns; we read these:

    Gene ID        → gene_candidates.gene (PK)
    Category       → gene_candidates.category
    Status         → gene_candidates.status (mapped via STATUS_MAP)
    Subcategory    → experiment_progress_logs row, message = "Subcategory: ..."
    Dec 2024 Upd.  → experiment_progress_logs row, timestamp = 2024-12-15
    Jun 2025 Upd.  → experiment_progress_logs row, timestamp = 2025-06-15
                     (including the literal "No Updates" string)

Discovery Scientist + Collaborating Scientist columns are read but not
written — emails are unknown, so user_email is left NULL. The UI
already handles NULL senders.

IDEMPOTENCY
-----------
- genes.upsert (PK = gene_id): no-op on existing.
- gene_candidates.upsert (PK = gene): no-op on existing. Existing
  curated rows are NOT overwritten with Excel data.
- experiment_progress_logs: each backfilled row carries
  {label: "backfill-2025-06", color: "#84cc16"}. Before inserting we
  SELECT logs for (gene) whose tags contain that label and skip any
  whose (message, timestamp) already exist. Safe to re-run.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from supabase import Client, create_client

# ---------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------

BACKFILL_TAG = {"label": "backfill-2025-06", "color": "#84cc16"}

# The gene_candidates_status_check constraint only allows these five
# values. The Excel uses different language; map each free-form label
# to the closest constraint-valid token. Unknown values fall back to
# 'suspected' (the lowest-confidence bucket) and emit a warning.
STATUS_MAP = {
    "Approved by TC": "in-translation",
    "Deprioritized": "stopped",
    "In progress": "under-investigation",
}
FALLBACK_STATUS = "suspected"

# Timestamps for the two date-named Excel columns. Use noon UTC so a
# DST shift won't flip the displayed day in the UI.
DEC_2024_TIMESTAMP = "2024-12-15T12:00:00+00:00"
JUN_2025_TIMESTAMP = "2025-06-15T12:00:00+00:00"

EXPECTED_HEADERS = {
    "Gene ID": 0,
    "Category": 1,
    "Status": 2,
    "Discovery Scientist": 3,
    "Collaborating Scientist": 4,
    "Subcategory": 5,
    "Dec 2024 Upd.": 6,
    "Jun 2025 Upd.": 7,
}


# ---------------------------------------------------------------------
# Data shape
# ---------------------------------------------------------------------


@dataclass
class GeneRow:
    """One row's worth of data ready to push to bloom."""

    gene_id: str
    category: str | None
    status: str
    progress_entries: list[dict]  # each: {message, timestamp}


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def normalize(v) -> str | None:
    """Excel cells come back as int/float/str/None — coerce to a
    trimmed str or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def map_status(raw: str | None) -> str:
    if raw is None:
        return FALLBACK_STATUS
    if raw in STATUS_MAP:
        return STATUS_MAP[raw]
    logging.warning("unknown status %r — using %s", raw, FALLBACK_STATUS)
    return FALLBACK_STATUS


def parse_rows(excel_path: Path) -> list[GeneRow]:
    """Read the Excel and emit one GeneRow per non-empty Gene ID row."""
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        sys.exit(f"Workbook {excel_path} has no active sheet")

    rows: list[GeneRow] = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        gene_id = normalize(raw[EXPECTED_HEADERS["Gene ID"]])
        if not gene_id:
            continue

        category = normalize(raw[EXPECTED_HEADERS["Category"]])
        status = map_status(normalize(raw[EXPECTED_HEADERS["Status"]]))

        progress: list[dict] = []
        subcategory = normalize(raw[EXPECTED_HEADERS["Subcategory"]])
        if subcategory:
            progress.append({"message": f"Subcategory: {subcategory}",
                             "timestamp": DEC_2024_TIMESTAMP})

        dec_upd = normalize(raw[EXPECTED_HEADERS["Dec 2024 Upd."]])
        if dec_upd:
            progress.append({"message": dec_upd,
                             "timestamp": DEC_2024_TIMESTAMP})

        jun_upd = normalize(raw[EXPECTED_HEADERS["Jun 2025 Upd."]])
        if jun_upd:
            progress.append({"message": jun_upd,
                             "timestamp": JUN_2025_TIMESTAMP})

        rows.append(GeneRow(gene_id=gene_id, category=category,
                            status=status, progress_entries=progress))

    return rows


# ---------------------------------------------------------------------
# Bloom writes (REST via supabase-py)
# ---------------------------------------------------------------------


def existing_backfill_logs(sb: Client, gene_id: str) -> set[tuple[str, str]]:
    """Return the set of (message, timestamp) pairs for this gene that
    were already backfilled (carry the BACKFILL_TAG). Used to skip
    re-inserts on re-runs.
    """
    res = (
        sb.table("experiment_progress_logs")
        .select("message,timestamp,tags")
        .eq("gene", gene_id)
        .execute()
    )
    seen: set[tuple[str, str]] = set()
    for row in res.data or []:
        tags = row.get("tags") or []
        if any(t.get("label") == BACKFILL_TAG["label"] for t in tags):
            seen.add((row["message"], row["timestamp"]))
    return seen


def push_row(sb: Client, row: GeneRow, dry_run: bool) -> tuple[int, int]:
    """Push one parsed row. Returns (gene_candidates_inserted,
    progress_logs_inserted) — 0 means skipped/existing."""
    g_inserted = 0
    l_inserted = 0

    if dry_run:
        logging.info("[dry-run] would upsert genes.gene_id=%s", row.gene_id)
        logging.info("[dry-run] would upsert gene_candidates(gene=%s, "
                     "category=%s, status=%s)",
                     row.gene_id, row.category, row.status)
        for p in row.progress_entries:
            logging.info("[dry-run] would insert progress log for %s @%s: %.80s",
                         row.gene_id, p["timestamp"], p["message"])
        return (1, len(row.progress_entries))

    # 1. seed genes (FK target for gene_candidates.gene)
    sb.table("genes").upsert(
        {"gene_id": row.gene_id}, on_conflict="gene_id"
    ).execute()

    # 2. upsert gene_candidates (no overwrite of existing curated rows)
    cand_res = sb.table("gene_candidates").upsert(
        {"gene": row.gene_id, "category": row.category, "status": row.status},
        on_conflict="gene",
        ignore_duplicates=True,
    ).execute()
    if cand_res.data:
        g_inserted = 1

    # 3. insert progress logs, skipping ones already backfilled
    seen = existing_backfill_logs(sb, row.gene_id)
    for p in row.progress_entries:
        if (p["message"], p["timestamp"]) in seen:
            continue
        sb.table("experiment_progress_logs").insert({
            "gene": row.gene_id,
            "message": p["message"],
            "timestamp": p["timestamp"],
            "user_email": None,
            "tags": [BACKFILL_TAG],
            "links": [],
            "images": [],
        }).execute()
        l_inserted += 1

    return (g_inserted, l_inserted)


# ---------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel-path", type=Path, required=True,
                    help="Path to HPI Gene Candidate Excel.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Log what would be sent; make no HTTP requests.")
    ap.add_argument("--limit", type=int, default=None,
                    help="Process only the first N rows (handy for spot-checks).")
    ap.add_argument("--verbose", action="store_true",
                    help="Debug-level logging.")
    args = ap.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    bloom_url = os.environ.get("BLOOM_URL")
    service_key = os.environ.get("SERVICE_ROLE_KEY")
    if not bloom_url or not service_key:
        sys.exit("BLOOM_URL and SERVICE_ROLE_KEY environment variables are required.")

    if not args.excel_path.exists():
        sys.exit(f"Excel file not found: {args.excel_path}")

    logging.info("Parsing %s", args.excel_path)
    rows = parse_rows(args.excel_path)
    if args.limit:
        rows = rows[: args.limit]
    logging.info("Parsed %d gene rows; %d total progress entries",
                 len(rows), sum(len(r.progress_entries) for r in rows))

    sb = create_client(bloom_url, service_key) if not args.dry_run else None
    if sb is None:
        logging.info("DRY RUN — no HTTP requests will be made.")

    total_genes = 0
    total_logs = 0
    for i, row in enumerate(rows, start=1):
        try:
            g, ll = push_row(sb, row, dry_run=args.dry_run)
        except Exception as e:  # noqa: BLE001 — surface and continue
            logging.error("row %d (gene=%s) failed: %s", i, row.gene_id, e)
            continue
        total_genes += g
        total_logs += ll
        if i % 25 == 0:
            logging.info("...processed %d/%d rows", i, len(rows))

    logging.info("")
    logging.info("Summary")
    logging.info("  gene_candidates inserted:  %d  (existing rows skipped)", total_genes)
    logging.info("  progress logs inserted:    %d", total_logs)
    if args.dry_run:
        logging.info("  (dry-run — nothing was sent over the wire)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
